from pathlib import Path
import base64
import os
import streamlit as st
import pandas as pd
from datetime import datetime

# ===================== CONFIGURAÇÃO: Excel local =====================
LOCAL_XLSX_PATH = r"C:\\Users\\AnaSilvaJeraCapital\\OneDrive - JERA CAPITAL GESTAO DE RECURSOS LTDA\\Comercial - Documentos\\NPS\\Pesquisa_NPS.xlsx"
SHOW_INTERNAL_NPS = False

# ===================== PÁGINA + CSS =====================
st.set_page_config(page_title="PESQUISA DE SATISFAÇÃO", layout="wide")

st.markdown(
    """
<style>
/* ===================== FONTES ===================== */

/* Ofelia Display Medium (Peso 500) */
@font-face {
  font-family: 'Ofelia Display';
  font-weight: 500;
  src: url('assets/fontes/OfeliaDisplay-Medium.ttf') format('truetype'),
       url('assets/fontes/OfeliaDisplay-Medium.woff2') format('woff2');
}

/* Ofelia Display Regular (Peso 400) */
@font-face {
  font-family: 'Ofelia Display';
  font-weight: 400;
  src: url('assets/fontes/OfeliaDisplay-Regular.ttf') format('truetype'),
       url('assets/fontes/OfeliaDisplay-Regular.woff2') format('woff2');
}

/* Ofelia Display Bold (Peso 700) */
@font-face {
  font-family: 'Ofelia Display';
  src: url('assets/fontes/OfeliaText-Bold.ttf') format('truetype');
  font-weight: 700;
}

/* Ofelia Text Regular (Peso 400) */
@font-face {
  font-family: 'Ofelia Text';
  font-weight: 400;
  src: url('assets/fontes/OfeliaText-Regular.otf') format('opentype'),
       url('assets/fontes/OfeliaText-Regular.ttf') format('truetype');
}

/* Ofelia Text Medium (Peso 500) */
@font-face {
  font-family: 'Ofelia Text';
  font-weight: 500;
  src: url('assets/fontes/OfeliaText-Medium.otf') format('opentype'),
       url('assets/fontes/OfeliaText-Medium.ttf') format('truetype');
}

/* Ofelia Text Semibold (Peso 650) */
@font-face {
  font-family: 'Ofelia Text';
  font-weight: 650;
  src: url('assets/fontes/OfeliaText-Semibold.ttf') format('truetype'),
       url('assets/fontes/OfeliaText-Semibold.woff2') format('woff2');
}

/* Ofelia Text Medium Italic (500 Italic) */
@font-face {
  font-family: 'Ofelia Text';
  src: url('assets/fontes/OfeliaText-MediumItalic.otf') format('opentype'),
       url('assets/fontes/OfeliaText-MediumItalic.ttf') format('truetype');
  font-weight: 500;
  font-style: italic;
}

/* ===================== VARIÁVEIS ===================== */
:root {
  --jera-primary:#00C1AD;
  --jera-dark:#052B38;
  --jera-bg:#052B38;
  --jera-light:#FFFFFF;
  --muted:#6b7c85;
}

/* ===================== RESET / BACKGROUND ===================== */
header[data-testid="stHeader"], footer {display:none !important;}
html, body, .stApp {
  background: var(--jera-bg) !important;
  font-family: 'Ofelia Text', sans-serif !important;
  color: var(--jera-dark);
  margin: 0 !important;
  padding: 0 !important;
  overflow-x: hidden !important;
}

/* ===================== CAIXA BRANCA ===================== */
div.block-container {
  background: rgba(255, 255, 255, 0.95) !important;
  border-radius: 22px !important;

  width: min(1200px, 96vw) !important;
  margin: 2vh auto !important;

  min-height: calc(100vh - 4vh) !important;
  height: auto !important;
  overflow-y: auto !important;

  padding: 3.2rem 4.0rem !important;
  box-shadow: 0 6px 18px rgba(0,0,0,.08);

  display: flex !important;
  flex-direction: column !important;
  align-items: center !important;
  box-sizing: border-box !important;
}

@media (max-width: 1200px){
  div.block-container { padding: 2.4rem 2.0rem !important; }
}

@media (max-width: 1024px){
  div.block-container {
    width: 100vw !important;
    margin: 0 auto !important;
    border-radius: 0 !important;
    min-height: 100vh !important;
    padding: 2.0rem 1.2rem 2.6rem 1.2rem !important;
  }
}

/* ===================== TIPOGRAFIA ===================== */
h1, h2, h3 {
  font-family: 'Ofelia Display', sans-serif !important;
  color: var(--jera-dark);
  text-align: center !important;
  margin-top: 0.4rem !important;
}

h1 { font-size: clamp(2.2rem, 3.2vw, 3.6rem) !important; font-weight: 700 !important; }
h2 { font-size: clamp(1.6rem, 2.3vw, 2.4rem) !important; font-weight: 650 !important; margin-bottom: 1.6rem !important; }
h3 { font-size: clamp(1.2rem, 1.8vw, 2.0rem) !important; font-weight: 550 !important; }

p, div, span, label {
  font-size: clamp(1.0rem, 1.2vw, 1.15rem) !important;
  line-height: 1.65 !important;
}

/* ===================== INPUT TELA 1 ===================== */
.codigo-cliente {
  font-family: 'Ofelia Display', sans-serif !important;
  font-size: 1.2rem !important;
  font-weight: 500 !important;
  text-align: center !important;
}

.pesquisa-identificada strong {
  font-family: 'Ofelia Display', sans-serif !important;
  font-weight: 700 !important;
}

p strong { font-weight: 650 !important; }

.stTextInput {
  display: flex !important;
  justify-content: center !important;
}
.stTextInput > div {
  width: fit-content !important;
  margin: 0 auto !important;
}
.stTextInput input {
  font-family: 'Ofelia Text', sans-serif !important;
  font-size: 1.05rem !important;
  text-align: center !important;
  padding: 0.65rem 0.9rem !important;
  border-radius: 10px !important;
  background-color: #f6f6f6 !important;
  width: 285px !important;
}
.stTextInput input::placeholder {
  color: #a0a0a0 !important;
  opacity: 1 !important;
}
div[data-testid="stTextInput"] > div > div:nth-child(1) > div:last-child {
  display: none !important;
}

/* ===================== BOTÕES ===================== */
.stButton > button {
  font-family: 'Ofelia Display', sans-serif !important;
  font-weight: 400 !important;
  font-size: 1.02rem !important;
  min-width: 220px !important;
  min-height: 48px !important;

  background: var(--jera-dark) !important;
  color: #ffffff !important;

  border: 1px solid rgba(5,43,56,0.70) !important;
  border-radius: 12px !important;

  box-shadow: 0 8px 18px rgba(0,0,0,.12) !important;
  transition: transform .18s ease, background .18s ease, box-shadow .18s ease, border-color .18s ease;
}

.stButton > button:hover {
  background: var(--jera-primary) !important;
  border-color: var(--jera-primary) !important;
  transform: translateY(-1px);
  box-shadow: 0 10px 22px rgba(0,0,0,.14) !important;
}

/* ===================== BLOCO DE CONTEÚDO ===================== */
.page {
  width: 100% !important;
  max-width: 1050px !important;
}

/* ===================== SLIDER / ESCALAS (1–5 MANTIDO) ===================== */
.scale-wrap {
  width: 100%;
  max-width: 760px;
  margin: 0.6rem auto 1.4rem auto;
}

.scale-ends {
  display:flex;
  justify-content: space-between;
  color: var(--muted);
  font-size: 0.95rem !important;
  margin-top: -0.35rem;
}

/* NUMERAÇÃO COMPLETA 1–5 */
.scale-numbers-5 {
  display: grid;
  grid-template-columns: repeat(5, 1fr);
  gap: 0;
  width: 100%;
  margin-top: -0.35rem;
  margin-bottom: 0.15rem;
  text-align: center;
  color: var(--muted);
  font-size: 0.95rem !important;
}
.scale-numbers-5 div {
  white-space: nowrap !important;
  text-align: center;
}
.scale-numbers-5 div:nth-child(1) { text-align: left; padding-left: 0.5rem; }
.scale-numbers-5 div:nth-child(5) { text-align: right; padding-right: 0.5rem; }

/* 2 alinhado com "Ruim" e 4 alinhado com "Bom" */
.scale-numbers-5 > div:nth-child(2) { transform: translateX(-1.5cm) !important; }
.scale-numbers-5 > div:nth-child(4) { transform: translateX( 1.5cm) !important; }

/* LABELS 1–5 (SEM NÚMERO) */
.scale-labels-5 {
  display: grid;
  grid-template-columns: repeat(5, 1fr);
  gap: 0;
  width: 100%;
  margin-top: 0.35rem;
  text-align: center;
}
.scale-labels-5 div { white-space: nowrap !important; font-size: 1.0rem !important; text-align: center; }

/* Péssimo */
.scale-labels-5 div:nth-child(1) { transform: translateX(-0.5cm); text-align: left; padding-left: 0.7rem; }
/* Ruim */
.scale-labels-5 div:nth-child(2) { transform: translateX(-1.5cm); }
/* Regular */
.scale-labels-5 div:nth-child(3) { transform: translateX(0); }
/* Bom */
.scale-labels-5 div:nth-child(4) { transform: translateX(1.5cm); }
/* Excelente */
.scale-labels-5 div:nth-child(5) { transform: translateX(0.5cm); text-align: right; padding-right: 0.7rem; }

/* ===================== ESTILO DO SLIDER ===================== */
div[data-testid="stSlider"] { width: 100% !important; }
div[data-testid="stSlider"] > div { padding-left: 0.75rem !important; padding-right: 0.75rem !important; }

div[data-testid="stSlider"] [data-baseweb="slider"] div[role="slider"] {
  border-color: var(--jera-primary) !important;
}
div[data-testid="stSlider"] [data-baseweb="slider"] div:nth-child(1) > div {
  background-color: var(--jera-primary) !important;
}

/* esconder numeração NATIVA do slider */
div[data-testid="stSlider"] [data-baseweb="slider"] span {
  display: none !important;
}

/* ===================== NOVO: NPS CUSTOM (SÓ USADO NO NPS) ===================== */
.nps-wrap{
  width: 100%;
  max-width: 860px;
  margin: 0.6rem auto 1.2rem auto;
  position: relative;
}

.nps-bar {
  height: 4px;
  background: rgba(0,193,173,0.35);
  border-radius: 999px;
  position: relative;
  margin: 0.7rem 0 0.85rem 0;
}

.nps-bar-fill{
  height: 4px;
  background: var(--jera-primary);
  border-radius: 999px;
  width: 0%;
}

.nps-knob{
  width: 18px;
  height: 18px;
  background: var(--jera-primary);
  border-radius: 999px;
  position: absolute;
  top: 50%;
  transform: translate(-50%,-50%);
  box-shadow: 0 6px 14px rgba(0,0,0,.18);
}

.nps-axis{
  position: relative;
  height: 28px;
}

.nps-tick{
  position: absolute;
  top: 0;
  transform: translateX(-50%);
  font-size: 0.95rem !important;
  color: var(--jera-dark);
  cursor: pointer;
  user-select: none;
  line-height: 1;
  padding: 2px 6px;
  border-radius: 10px;
}

.nps-tick.selected{
  color: #fff !important;
  background: var(--jera-primary);
  font-weight: 700 !important;
}

.nps-tick:hover{
  background: rgba(0,193,173,0.15);
}

/* Slider do NPS invisível (mas funcional) */
.nps-hidden-slider{
  position:absolute;
  left:-99999px;
  top:-99999px;
  width:1px;
  height:1px;
  opacity:0;
  pointer-events:none;
}

/* ===================== RODAPÉ FIXO ===================== */
.footer-fixed {
  position: fixed !important;
  bottom: 10px !important;
  left: 16px !important;
  font-size: 0.9rem !important;
  color: #7A8C94 !important;
  font-family: 'Ofelia Text', sans-serif !important;
  z-index: 9999 !important;
  pointer-events: none;
}

/* ===================== TELA 1: AJUSTES DO TÍTULO ===================== */
.tela-1 .h1-tela1{
  margin-top: -70px !important;
  margin-bottom: 0.5rem !important;
}

/* ===================== INSTRUÇÃO DE SLIDER ===================== */
.slider-instruction {
  text-align: center;
  color: #00C1AD;
  font-weight: 600;
  margin-top: -1.0rem;
  margin-bottom: 0.5rem;
  font-size: 1.0rem !important;
}
</style>
""",
    unsafe_allow_html=True,
)

# ===================== LOGO =====================
BASE_DIR = Path(__file__).parent.resolve()
ASSETS = BASE_DIR / "assets"
LOGO_FULL = ASSETS / "jera-logo-full.png"

def _img_data_uri(p: Path) -> str:
    return "data:image/png;base64," + base64.b64encode(p.read_bytes()).decode()

# ===================== ESTADO INICIAL =====================
if "step" not in st.session_state:
    st.session_state["step"] = 1
if "client_code" not in st.session_state:
    st.session_state["client_code"] = ""

# flags de “mexeu no slider”
def _touch(key: str):
    st.session_state[f"{key}__touched"] = True

# ===================== BLOCOS DE PERGUNTAS =====================
BLOCOS = [
    (
        "Qualidade do Relacionamento com a Equipe Jera",
        [
            (
                "Tempo de resolução às solicitações",
                "De 01 a 05, quanto você está satisfeito(a) com a agilidade e disponibilidade da equipe ao atender suas solicitações?",
            ),
            (
                "Proatividade na comunicação",
                "De 01 a 05, quanto a equipe se antecipa às suas necessidades e se comunica de forma proativa?",
            ),
        ],
    ),
    (
        "Clareza e Relevância das Informações Prestadas",
        [
            (
                "Clareza das informações apresentadas",
                "De 01 a 05, o quanto as informações e o detalhamento dos relatórios atendem às suas expectativas?",
            ),
            (
                "Compreensão dos resultados",
                "De 01 a 05, o quanto os relatórios ajudam você a entender se a carteira está caminhando conforme seus objetivos?",
            ),
        ],
    ),
    (
        "Efetividade dos Encontros e Alinhamentos",
        [
            (
                "Frequência, formato e duração das reuniões",
                "De 01 a 05, como você avalia a adequação da frequência, do formato e da duração das reuniões?",
            ),
            (
                "Relevância e efetividade das reuniões",
                "De 01 a 05, o quanto as reuniões apresentam conteúdos relevantes, claros e bem organizados?",
            ),
        ],
    ),
    (
        "Percepção sobre o Desempenho da Carteira",
        [
            (
                "Satisfação com o retorno obtido",
                "De 01 a 05, o quanto você está satisfeito com o retorno da sua carteira nos últimos meses?",
            ),
            (
                "Alinhamento entre retorno e perfil de risco",
                "De 01 a 05, o quanto o retorno da carteira está compatível com seu perfil de risco e objetivos financeiros?",
            ),
        ],
    ),
    (
        "Compromisso com a Transparência e Integridade",
        [
            (
                "Independência nas recomendações",
                "De 01 a 05, o quanto você percebe independência e isenção nas recomendações feitas pela equipe?",
            ),
            (
                "Transparência sobre custos e remunerações",
                "De 01 a 05, o quanto você sente clareza nas informações sobre custos, taxas e formas de remuneração?",
            ),
        ],
    ),
]

HEADERS = (
    ["timestamp", "client_code", "NPS"]
    + [p[0] for _, perguntas in BLOCOS for p in perguntas]
    + ["coment_final"]
)

# ===================== FUNÇÕES AUXILIARES =====================
def _append_to_excel(row_values):
    try:
        from openpyxl import Workbook, load_workbook

        os.makedirs(os.path.dirname(LOCAL_XLSX_PATH), exist_ok=True)

        if os.path.exists(LOCAL_XLSX_PATH):
            wb = load_workbook(LOCAL_XLSX_PATH)
        else:
            wb = Workbook()
            if wb.active and wb.active.title != "Respostas":
                wb.remove(wb.active)

        ws = wb["Respostas"] if "Respostas" in wb.sheetnames else wb.create_sheet("Respostas")

        for col, header in enumerate(HEADERS, 1):
            ws.cell(row=1, column=col, value=header)

        next_row = ws.max_row + 1
        for col, val in enumerate(row_values, 1):
            ws.cell(row=next_row, column=col, value=val)

        wb.save(LOCAL_XLSX_PATH)
        return True, "Gravado no Excel local."
    except Exception as e:
        return False, str(e)

def escala_1a5(key: str) -> int:
    if f"{key}__touched" not in st.session_state:
        st.session_state[f"{key}__touched"] = False

    if key not in st.session_state:
        st.session_state[key] = 1

    if not st.session_state.get(f"{key}__touched", False):
        st.markdown("<p class='slider-instruction'>Deslize para responder</p>", unsafe_allow_html=True)
    else:
        st.markdown("<div style='height: 1.0rem;'></div>", unsafe_allow_html=True)

    st.markdown("<div class='scale-wrap'>", unsafe_allow_html=True)
    val = st.slider(
        label="",
        min_value=1,
        max_value=5,
        value=st.session_state[key],
        step=1,
        key=key,
        on_change=_touch,
        args=(key,),
        label_visibility="collapsed",
    )

    st.markdown(
        """
        <div class="scale-numbers-5">
          <div>1</div><div>2</div><div>3</div><div>4</div><div>5</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        <div class="scale-labels-5">
          <div>Péssimo</div>
          <div>Ruim</div>
          <div>Regular</div>
          <div>Bom</div>
          <div>Excelente</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("</div>", unsafe_allow_html=True)
    return val

# ======= NPS (REFATORADO APENAS AQUI) =======
def escala_0a10(key: str) -> int:
    if f"{key}__touched" not in st.session_state:
        st.session_state[f"{key}__touched"] = False

    if key not in st.session_state:
        st.session_state[key] = 0

    if not st.session_state.get(f"{key}__touched", False):
        st.markdown("<p class='slider-instruction'>Deslize para responder</p>", unsafe_allow_html=True)
    else:
        st.markdown("<div style='height: 1.0rem;'></div>", unsafe_allow_html=True)

    # Slider invisível (somente para manter o valor e permitir arrastar via teclado/acessibilidade)
    st.markdown("<div class='nps-hidden-slider'>", unsafe_allow_html=True)
    val = st.slider(
        label="",
        min_value=0,
        max_value=10,
        value=st.session_state[key],
        step=1,
        key=key,
        on_change=_touch,
        args=(key,),
        label_visibility="collapsed",
    )
    st.markdown("</div>", unsafe_allow_html=True)

    # UI custom do NPS
    pct = (val / 10) * 100.0

    ticks_html = []
    for n in range(0, 11):
        left = (n / 10) * 100.0
        cls = "nps-tick selected" if n == val else "nps-tick"
        ticks_html.append(f"<span class='{cls}' style='left:{left:.4f}%' data-n='{n}'>{n}</span>")
    ticks_html = "\n".join(ticks_html)

    st.components.v1.html(
        f"""
        <div class="nps-wrap" id="nps-wrap">
          <div class="nps-bar">
            <div class="nps-bar-fill" style="width:{pct:.4f}%"></div>
            <div class="nps-knob" style="left:{pct:.4f}%"></div>
          </div>
          <div class="nps-axis">
            {ticks_html}
          </div>
        </div>

        <script>
          (function() {{
            const root = document.currentScript.parentElement;
            const ticks = root.querySelectorAll('.nps-tick');

            // tenta achar o input range do Streamlit (o slider invisível ainda existe no DOM)
            function findSliderInput() {{
              // pega o primeiro input[type="range"] anterior no DOM
              const ranges = window.parent.document.querySelectorAll('input[type="range"]');
              // usa o último, normalmente é o mais recente renderizado
              return ranges[ranges.length - 1] || null;
            }}

            const slider = findSliderInput();

            function setValue(n) {{
              if (!slider) return;
              slider.value = n;
              slider.dispatchEvent(new Event('input', {{ bubbles: true }}));
              slider.dispatchEvent(new Event('change', {{ bubbles: true }}));
            }}

            ticks.forEach(t => {{
              t.addEventListener('click', () => {{
                const n = parseInt(t.getAttribute('data-n'));
                setValue(n);
              }});
            }});
          }})();
        </script>
        """,
        height=120,
    )

    return val

# ===================== FLUXO DAS TELAS =====================
step = st.session_state["step"]
st.markdown("<div class='page'>", unsafe_allow_html=True)

# -------- TELA 1 --------
if step == 1:
    st.markdown("<div class='tela-1'>", unsafe_allow_html=True)

    if LOGO_FULL.exists():
        st.markdown(
            f"<img alt='Jera' src='{_img_data_uri(LOGO_FULL)}' "
            "style='display:block;margin:-90px auto -15px auto;width:480px;max-width:95%;'/>",
            unsafe_allow_html=True,
        )

    st.markdown(
        """
        <h1 class="h1-tela1" style="font-size: 2.0rem; line-height: 1; transform: translateX(0.6cm);">
          PESQUISA DE SATISFAÇÃO
        </h1>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("<div style='height:1.2rem;'></div>", unsafe_allow_html=True)

    st.markdown("<p class='codigo-cliente'><strong>CÓDIGO DO CLIENTE</strong></p>", unsafe_allow_html=True)
    st.text_input("", key="client_code", placeholder="Ex.: APELIDO", max_chars=20)

    st.markdown("<div style='height:0.8rem;'></div>", unsafe_allow_html=True)

    st.markdown(
        """
        <div style='text-align:center; line-height:1.6; margin-bottom:0.6rem;'>
          <p class='pesquisa-identificada' style='margin-bottom:0.4rem;'><strong>Esta é uma pesquisa identificada.</strong></p>
          <p style='font-size:1.05rem; margin-top:0; font-style: italic; font-weight: 500;'>
            Suas respostas serão tratadas com confidencialidade e utilizadas exclusivamente
            para aperfeiçoarmos nossos serviços, sempre alinhados aos seus objetivos.
          </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("<div style='height:0.6rem;'></div>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1.4, 1, 1])
    with col2:
        if st.button("Iniciar pesquisa", key="start_button"):
            if not st.session_state["client_code"].strip():
                st.error("Por favor, preencha o código do cliente.")
            else:
                st.session_state["step"] = 2
                st.rerun()

    st.markdown("</div>", unsafe_allow_html=True)

# -------- TELAS 2–6 --------
elif 2 <= step <= 6:
    idx = step - 2
    titulo, perguntas = BLOCOS[idx]

    st.markdown(f"<h2>{titulo}</h2>", unsafe_allow_html=True)
    st.markdown("<div style='height:0.5rem;'></div>", unsafe_allow_html=True)

    respostas = st.session_state.get(f"respostas_{idx}", {})

    for i, (topico, texto) in enumerate(perguntas):
        st.markdown(f"<h3 style='margin-bottom:0.2rem;'>{topico}</h3>", unsafe_allow_html=True)
        st.markdown(f"<p style='text-align:center; margin-top:0;'>{texto}</p>", unsafe_allow_html=True)

        pergunta_key = f"{titulo}__{i}"
        val = escala_1a5(pergunta_key)
        respostas[topico] = val

        st.markdown("<div style='height:1.1rem;'></div>", unsafe_allow_html=True)

    st.session_state[f"respostas_{idx}"] = respostas

    col1, col2, col3 = st.columns([2, 6, 2])
    with col1:
        if st.button("◀ Voltar"):
            st.session_state["step"] -= 1
            st.rerun()
    with col3:
        if st.button("Avançar ►"):
            st.session_state["step"] += 1
            st.rerun()

# -------- PÁGINA NPS --------
elif step == 7:
    st.markdown("<h2>NPS</h2>", unsafe_allow_html=True)
    st.markdown(
        """
        <p style='line-height:1.55; margin-bottom:1.0rem; text-align:center;'>
        Considerando sua experiência com os serviços da <b>Jera Capital</b> ao longo do último ano — incluindo
        atendimento, relatórios, reuniões, transparência e a adequação das soluções ao seu perfil —,
        em uma escala de <b>0 a 10</b>, o quanto você recomendaria a Jera Capital a amigos ou familiares?
        </p>
        <p style='text-align:center; color:#334a55; margin-top:0;'>
          <em>(0 = Não recomendaria de forma alguma | 10 = Recomendaria com total confiança)</em>
        </p>
        """,
        unsafe_allow_html=True,
    )

    nps = escala_0a10("nps_score")
    st.markdown("<div style='height:1.2rem;'></div>", unsafe_allow_html=True)

    st.markdown(
        """
        <p style='font-weight:750; margin-bottom:0.2rem; text-align:left; width:100%; max-width:900px;'>
          Comentário final:
        </p>
        <p style='margin-top:0; color:#334a55; text-align:left; width:100%; max-width:900px;'>
          Se desejar, utilize este espaço para compartilhar sugestões, elogios ou qualquer ponto que não tenha sido abordado anteriormente.
        </p>
        """,
        unsafe_allow_html=True,
    )

    coment_final = st.text_area("", placeholder="", key="coment_final")

    col1, col2, col3 = st.columns([2, 6, 2])
    with col1:
        if st.button("◀ Voltar"):
            st.session_state["step"] -= 1
            st.rerun()

    with col3:
        if st.button("Enviar respostas ✅"):
            code = st.session_state["client_code"].strip()
            if not code:
                st.error("O campo CÓDIGO DO CLIENTE é obrigatório.")
                st.stop()

            row = {
                "timestamp": datetime.now().isoformat(timespec="seconds"),
                "client_code": code,
                "NPS": nps,
            }

            for i, (_, perguntas) in enumerate(BLOCOS):
                respostas = st.session_state.get(f"respostas_{i}", {})
                for topico, _ in perguntas:
                    row[topico] = respostas.get(topico)

            row["coment_final"] = coment_final

            try:
                df_old = pd.read_csv("responses.csv")
                df = pd.concat([df_old, pd.DataFrame([row])], ignore_index=True)
            except FileNotFoundError:
                df = pd.DataFrame([row])

            df.to_csv("responses.csv", index=False)
            ok, _msg = _append_to_excel([row.get(h) for h in HEADERS])

            if ok:
                st.success("Respostas gravadas com sucesso no Excel! ✔")
            else:
                st.warning(f"Não foi possível gravar no Excel. As respostas foram salvas em responses.csv. (Erro: {_msg})")

            st.session_state["step"] = 8
            st.rerun()

# -------- CONFIRMAÇÃO FINAL --------
elif step == 8:
    st.markdown("<h2>✅ Resposta enviada com sucesso</h2>", unsafe_allow_html=True)
    st.success(
        "Agradecemos por dedicar seu tempo para responder à nossa pesquisa. "
        "Suas respostas são muito importantes para que possamos aprimorar continuamente "
        "a qualidade dos nossos serviços e o relacionamento com você."
    )

    st.caption(
        f"Código do cliente: **{st.session_state['client_code']}** • "
        f"Enviado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )

    st.markdown(
        "<p style='margin-top:1.2rem;'>"
        "Caso tenha qualquer dúvida ou queira conversar conosco, nossa equipe está sempre à disposição."
        "</p>",
        unsafe_allow_html=True,
    )

    if st.button("➕ Enviar nova resposta"):
        for k in list(st.session_state.keys()):
            if k.startswith("respostas_") or k in ["nps_score", "coment_final"]:
                st.session_state.pop(k, None)
            if k.endswith("__touched"):
                st.session_state.pop(k, None)

        st.session_state["client_code"] = ""
        st.session_state["step"] = 1
        st.rerun()

st.markdown("</div>", unsafe_allow_html=True)

# -------- RODAPÉ FIXO --------
st.markdown("<div class='footer-fixed'>© Jera Capital — Todos os direitos reservados.</div>", unsafe_allow_html=True)
